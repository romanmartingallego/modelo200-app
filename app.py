import streamlit as st
import pdfplumber
import pandas as pd
import re
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from io import BytesIO

st.set_page_config(page_title="Modelo 200", page_icon="ğŸ“„")
st.title("ğŸ“„ Procesador de Modelo 200")
st.write("Sube tus archivos PDF y una plantilla Excel. El sistema rellenarÃ¡ automÃ¡ticamente los datos en la plantilla y te permitirÃ¡ descargar el resultado.")

# âœ… Inicializar clave para reinicio de file_uploader
if "upload_key" not in st.session_state:
    st.session_state.upload_key = 0

# ğŸ” BotÃ³n de reinicio total
if st.button("ğŸ”„ Reiniciar formulario"):
    st.session_state.clear()
    st.session_state.upload_key += 1
    st.rerun()

# ğŸ“‚ Subida de archivos controlada por clave dinÃ¡mica
pdf_files = st.file_uploader(
    "ğŸ”¼ Sube uno o mÃ¡s archivos PDF del Modelo 200",
    type="pdf",
    accept_multiple_files=True,
    key=f"pdf_uploader_{st.session_state.upload_key}"
)

excel_file = st.file_uploader(
    "ğŸ“Š Sube la plantilla Excel",
    type="xlsx",
    key=f"excel_uploader_{st.session_state.upload_key}"
)

# âœ… Guardar archivos en la sesiÃ³n
if pdf_files:
    st.session_state.pdfs = pdf_files
if excel_file:
    st.session_state.excel = excel_file

# âœ… Procesamiento solo si todo estÃ¡ subido y se pulsa el botÃ³n
if "pdfs" in st.session_state and "excel" in st.session_state:
    if st.session_state.pdfs and st.session_state.excel:
        st.success("âœ… Archivos cargados correctamente.")
        if st.button("ğŸš€ Procesar archivos"):
            uploaded_pdfs = st.session_state.pdfs
            uploaded_excel = st.session_state.excel

            excel_bytes = BytesIO(uploaded_excel.read())
            workbook = load_workbook(excel_bytes)
            sheet = workbook["Modelo 200 input"]

            df_full = pd.read_excel(excel_bytes, sheet_name="Modelo 200 input", header=None)
            headers_row = df_full.iloc[9]

            columnas_ano = {
                int(valor): idx
                for idx, valor in headers_row.items()
                if pd.notna(valor) and isinstance(valor, (int, float)) and 2000 <= int(valor) <= 2100
            }

            codigos_en_plantilla = {
                str(fila[0].value).strip().zfill(5): fila[0].row
                for fila in sheet.iter_rows(min_row=11, min_col=3, max_col=3)
                if fila[0].value
            }

            def extraer_ano(pdf_bytes):
                with pdfplumber.open(pdf_bytes) as pdf:
                    for i, pagina in enumerate(pdf.pages):
                        texto = pagina.extract_text()
                        if not texto:
                            continue
                        match = re.search(r'\b(20[1-2][0-9])\d{11}[A-Z]?\b', texto)
                        if match:
                            return int(match.group(1))
                return None

            def extraer_codigos_valores(pdf_bytes):
                resultados = []
                patron = re.compile(r'(\d{5})\s+([-]?\d{1,3}(?:\.\d{3})*,\d{2})')
                SECCIONES_RELEVANTES = [
                    "Balance: Activo",
                    "Balance: Patrimonio neto y pasivo",
                    "Cuenta de pÃ©rdidas y ganancias"
                ]
                with pdfplumber.open(pdf_bytes) as pdf:
                    for pagina in pdf.pages:
                        texto = pagina.extract_text()
                        if not texto:
                            continue
                        if not any(seccion in texto for seccion in SECCIONES_RELEVANTES):
                            continue
                        coincidencias = patron.findall(texto)
                        for codigo, valor in coincidencias:
                            valor_num = float(valor.replace('.', '').replace(',', '.'))
                            resultados.append((codigo, valor_num))
                return resultados

            for pdf in uploaded_pdfs:
                pdf_bytes = BytesIO(pdf.read())
                aÃ±o_fiscal = extraer_ano(pdf_bytes)
                if not aÃ±o_fiscal:
                    st.warning(f"âŒ No se pudo detectar el aÃ±o fiscal en {pdf.name}")
                    continue
                if aÃ±o_fiscal not in columnas_ano:
                    st.warning(f"âš ï¸ El aÃ±o {aÃ±o_fiscal} no estÃ¡ en la plantilla. Saltando {pdf.name}")
                    continue

                pdf_bytes.seek(0)
                datos = extraer_codigos_valores(pdf_bytes)

                col_idx = columnas_ano[aÃ±o_fiscal]
                col_letter = get_column_letter(col_idx + 1)

                for codigo, fila in codigos_en_plantilla.items():
                    sheet[f"{col_letter}{fila}"] = None

                encontrados = 0
                for codigo, valor in datos:
                    codigo_formateado = str(codigo).strip().zfill(5)
                    if codigo_formateado in codigos_en_plantilla:
                        fila = codigos_en_plantilla[codigo_formateado]
                        sheet[f"{col_letter}{fila}"] = valor
                        encontrados += 1

                st.success(f"âœ… {pdf.name} procesado correctamente ({encontrados} valores escritos en {aÃ±o_fiscal})")

            output = BytesIO()
            workbook.save(output)
            output.seek(0)

            st.download_button(
                label="ğŸ“¥ Descargar Excel Modificado",
                data=output,
                file_name="Modelo_200_completo.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
